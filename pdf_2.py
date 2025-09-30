"""
PDF Table Extractor with advanced header analysis to handle all table structures,
prevent duplicate headers, and correctly process all data.
"""

import camelot
import pandas as pd
import os
import re
import fitz  # PyMuPDF
from typing import Optional, Tuple, List, Dict
from openpyxl.utils import get_column_letter
import logging
from datetime import datetime

# Setup logging to capture only our script's output, not library debug info
log_filename = f"pdf_extraction_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

# Set up our custom logger
logger = logging.getLogger('pdf_extractor')
logger.setLevel(logging.INFO)

# Create formatter
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

# File handler for our log
file_handler = logging.FileHandler(log_filename, encoding='utf-8')
file_handler.setLevel(logging.INFO)
file_handler.setFormatter(formatter)

# Console handler
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_handler.setFormatter(formatter)

# Add handlers to our logger
logger.addHandler(file_handler)
logger.addHandler(console_handler)

# Disable debug logging from other libraries
logging.getLogger('camelot').setLevel(logging.WARNING)
logging.getLogger('pdfminer').setLevel(logging.WARNING)
logging.getLogger('fitz').setLevel(logging.WARNING)
logging.getLogger('pypdf').setLevel(logging.WARNING)

def log_print(*args, **kwargs):
    """Custom print function that logs to both console and file"""
    message = ' '.join(str(arg) for arg in args)
    logger.info(message)
    # Also print to console normally
    print(message)

# Don't replace the built-in print globally, use log_print where needed

# --- Helper Functions ---

def _sanitize_text(text: str) -> str:
    """Cleans up text to be used as a filename or key."""
    text = text.strip()
    text = re.sub(r"\s+", " ", text)
    return text[:160]

def _to_fitz_y(page, camelot_y: float) -> float:
    """Converts Camelot's y-coordinate to PyMuPDF's y-coordinate."""
    return float(page.rect.height) - float(camelot_y)

def _horizontal_overlap(bbox1, bbox2):
    """Calculates the horizontal overlap ratio of two bounding boxes."""
    x1_l, _, x1_r, _ = bbox1
    x2_l, _, x2_r, _ = bbox2
    if x1_r < x2_l or x2_r < x1_l: return 0
    intersection_width = min(x1_r, x2_r) - max(x1_l, x2_l)
    width1 = x1_r - x1_l
    width2 = x2_r - x2_l
    return intersection_width / min(width1, width2) if min(width1, width2) > 0 else 0

def _find_heading_above(page, bbox, band_px: int = 120) -> Optional[str]:
    """Finds a likely heading text block just above a table's bounding box."""
    try: _, _, _, y_top = bbox
    except (TypeError, ValueError): return None
    table_top_fitz = _to_fitz_y(page, y_top)
    blocks = page.get_text("blocks") or []
    candidates = []
    for b in blocks:
        try: x0, y0, x1, y1, text, *_ = b
        except ValueError: continue
        if not text or not text.strip(): continue
        if y1 <= table_top_fitz and (table_top_fitz - y1) <= band_px:
            s = text.strip()
            if 2 <= len(s) <= 140:
                digits = sum(c.isdigit() for c in s)
                letters = sum(c.isalpha() for c in s)
                if letters == 0 and digits > 0: continue
                if digits / max(1, len(s)) > 0.4: continue
                distance = table_top_fitz - y1
                candidates.append((distance, s))
    if not candidates: return None
    candidates.sort(key=lambda t: t[0])
    return _sanitize_text(candidates[0][1])

# --- Core Extraction and Grouping Logic ---

def extract_and_group_tables(pdf_path, pages='all', min_rows=1, min_cols=1, debug=False):
    """Extracts tables with Camelot and then groups them across pages."""
    try:
        if not os.path.exists(pdf_path):
            raise FileNotFoundError(f"PDF file not found: {pdf_path}")
        log_print(f"Extracting tables from: {pdf_path} using Camelot...")
        tables = camelot.read_pdf(pdf_path, pages=pages, flavor='lattice', line_scale=40)
        if not tables:
            log_print("No tables found by Camelot."); return {}
        log_print(f"Found {len(tables)} tables. Filtering and grouping...")
        filtered_tables = [t for t in tables if len(t.df) >= min_rows and len(t.df.columns) >= min_cols]

        def safe_bbox(t):
            return getattr(t, "_bbox", t.parsing_report.get("bbox", (0, 0, 0, 0)))

        sorted_filtered_tables = sorted(filtered_tables, key=lambda t: (int(t.page), -safe_bbox(t)[3], safe_bbox(t)[0]))
        groups = []
        doc = fitz.open(pdf_path)

        for idx, t in enumerate(sorted_filtered_tables, start=1):
            current_page = int(t.page)
            page = doc[current_page - 1]
            bbox = safe_bbox(t)
            heading = _find_heading_above(page, bbox, band_px=140)
            t.heading = heading
            if debug: log_print(f"\n[Table {idx}] Page {current_page}, BBox={bbox}\n  Detected heading: {heading if heading else 'None'}")
            if groups:
                last_group = groups[-1]
                last_table_in_group = last_group[-1]
                last_page = int(last_table_in_group.page)
                if heading: groups.append([t])
                else:
                    overlap = _horizontal_overlap(bbox, safe_bbox(last_table_in_group))
                    if current_page == last_page + 1 and overlap > 0.5: last_group.append(t)
                    else: groups.append([t])
            else: groups.append([t])

        final_tables = {}
        for i, group in enumerate(groups):
            merged_df = pd.concat([t.df for t in group], ignore_index=True)
            group_heading = group[0].heading
            key = group_heading if group_heading else f"Table_Group_{i+1}Page{group[0].page}"
            original_key, counter = key, 1
            while key in final_tables:
                key = f"{original_key}_{counter}"; counter += 1
            final_tables[key] = merged_df
        doc.close()
        for h, d in final_tables.items(): log_print(f"\nGroup: '{h}' -> Rows: {len(d)}, Cols: {len(d.columns)}")
        return final_tables
    except Exception as e:
        log_print(f"Error extracting tables: {str(e)}"); return {}

# --- Function to Save Raw Tables for Debugging ---

def save_raw_tables_for_debug(tables_dict, output_filename="debug_raw_tables.xlsx"):
    """Saves each extracted DataFrame to a separate sheet in an Excel file for inspection."""
    log_print(f"\nSaving raw tables for debugging to '{output_filename}'...")
    try:
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            for table_name, df in tables_dict.items():
                safe_sheet_name = re.sub(r'[\\/*?:"<>|]', "", table_name)[:31]
                df.to_excel(writer, sheet_name=safe_sheet_name, index=False, header=False)
        log_print("✅ Debug file saved successfully.")
    except Exception as e:
        log_print(f"❌ Could not save debug file: {e}")

# --- ADVANCED HEADER ANALYSIS FUNCTION ---

def find_header_and_data_start(df: pd.DataFrame) -> Tuple[int, int, list]:
    """
    Analyzes the first few rows of a DataFrame to find the true header row,
    data start row, and constructs a merged, clean header.
    """
    def is_row_numerical_index(row):
        return all(str(item).strip() == str(i) for i, item in enumerate(row))
    
    def has_year_pattern(row):
        """Check if row contains year patterns like 2018-19, 2019-20, etc."""
        year_pattern = re.compile(r'\b20\d{2}-\d{2}\b')
        return any(year_pattern.search(str(cell)) for cell in row if pd.notna(cell))
    
    def count_meaningful_cells(row):
        """Count cells with meaningful content (not empty, not just numbers)"""
        meaningful = 0
        for cell in row:
            cell_str = str(cell).strip()
            if cell_str and cell_str != 'nan' and len(cell_str) > 1:
                # Check if it's not just a single number
                if not (cell_str.isdigit() and len(cell_str) <= 3):
                    meaningful += 1
        return meaningful
    
    def is_header_row(row):
        """Check if a row looks like a header row"""
        row_str = [str(cell).strip().lower() for cell in row]
        # Look for common header indicators
        header_keywords = ['academic year', 'program', 'students', 'year', 'total', 'no.', 'amount']
        return any(any(keyword in cell for keyword in header_keywords) for cell in row_str)
    
    start_offset = 0
    if len(df) > 0 and is_row_numerical_index(df.iloc[0]):
        start_offset = 1

    best_header_index = -1
    
    # Strategy 1: Look for obvious header rows first
    search_range = min(start_offset + 4, len(df))
    for i in range(start_offset, search_range):
        if i >= len(df):
            break
        row = df.iloc[i]
        
        # Skip rows that are clearly data (contain year patterns but no header keywords)
        if has_year_pattern(row) and not is_header_row(row):
            continue
            
        # Check if this looks like a header row
        if is_header_row(row):
            best_header_index = i
            break
    
    # Strategy 2: If no obvious header found, use meaningful content approach
    if best_header_index == -1:
        max_meaningful_cells = -1
        for i in range(start_offset, search_range):
            if i >= len(df):
                break
            row = df.iloc[i]
            meaningful_cells = count_meaningful_cells(row)
            
            # Skip rows that look like data (contain years)
            if has_year_pattern(row):
                continue
                
            # Prefer rows with more meaningful content
            if meaningful_cells > max_meaningful_cells:
                max_meaningful_cells = meaningful_cells
                best_header_index = i

    # Strategy 3: Final fallback
    if best_header_index == -1:
        best_header_index = start_offset if start_offset < len(df) else 0

    # Find where data actually starts 
    data_start_index = best_header_index + 1
    
    # Don't try to merge headers for simple cases - just use the header row as is
    if best_header_index < len(df):
        primary_header = [str(h).replace('\n', ' ').strip() for h in df.iloc[best_header_index]]
    else:
        primary_header = [f"Column_{i}" for i in range(len(df.columns))]
    
    # Only merge if we have a complex table structure with multiple potential header rows
    # and the primary header doesn't already look complete
    should_merge = False
    if best_header_index > start_offset:
        # Check if primary header has empty or incomplete cells
        empty_or_short = sum(1 for h in primary_header if not h or h == 'nan' or len(h.strip()) < 2)
        if empty_or_short > len(primary_header) * 0.3:  # More than 30% empty/short
            should_merge = True
    
    if should_merge:
        # Merge any info from rows above the primary header
        for i in range(start_offset, best_header_index):
            if i < len(df):
                secondary_row = [str(h).replace('\n', ' ').strip() for h in df.iloc[i]]
                for j, text in enumerate(secondary_row):
                    if j < len(primary_header) and text and text not in primary_header[j] and len(text) > 2:
                        if primary_header[j] and primary_header[j] != 'nan':
                            primary_header[j] = f"{text} {primary_header[j]}".strip()
                        else:
                            primary_header[j] = text
    
    # Clean up header names
    cleaned_header = []
    for header in primary_header:
        if header and header != 'nan' and header.strip():
            cleaned_header.append(header.strip())
        else:
            cleaned_header.append(f"Column_{len(cleaned_header)}")
                
    return best_header_index, data_start_index, cleaned_header

def process_pivot_table(df_data, headers, table_heading, debug=False):
    """Block-based processing of repeating Academic Year sections.

    Each occurrence of 'Academic Year' (or its auto-suffixed variant) starts a new block.
    All columns until (but excluding) the next 'Academic Year' column belong to that block.
    For every row, the year value appearing under the block's own Academic Year column is paired
    ONLY with the metrics inside that same block. This prevents accidental pairing of a metric
    with an earlier/later year's value simply because they share a physical row.

    Returns a list of record dicts: {'Header': (table_heading, metric_name, academic_year), 'Value': value}
    The outer pipeline may later drop the table_heading level for the final 2-level (Metric, Year) layout.
    """
    if debug:
        print("Block-based pivot normalization")
        print(f"Headers: {headers}")
        print(f"Shape: {df_data.shape}")

    # Identify positions of Academic Year columns
    year_positions = [i for i, h in enumerate(headers) if h == 'Academic Year' or h.startswith('Academic Year_')]
    if not year_positions:
        if debug:
            print("No repeating Academic Year columns found; fallback to generic handling.")
        return []

    # Append sentinel (end of columns) for block slicing logic
    year_positions_sorted = sorted(year_positions)
    year_positions_sorted.append(len(headers))

    table_heading_clean = str(table_heading).replace('\n', ' ').strip()
    processed_data: List[Dict] = []

    # Iterate over blocks
    for idx in range(len(year_positions_sorted) - 1):
        year_col = year_positions_sorted[idx]
        next_boundary = year_positions_sorted[idx + 1]
        metric_cols = [c for c in range(year_col + 1, next_boundary) if c < len(headers)]
        block_id = idx + 1
        if debug:
            block_header_names = [headers[c] for c in metric_cols]
            print(f"Block {block_id}: year_col={year_col} -> metrics={block_header_names}")

        for row_i, row in df_data.iterrows():
            # Extract the year for this block
            if year_col >= len(row):
                continue
            year_raw = str(row.iloc[year_col]).strip()
            if not year_raw or year_raw.lower() in ['nan', '-']:
                # Skip rows without a valid year in this block
                continue

            # For each metric in this block, pair with the block's year only
            for mc in metric_cols:
                if mc >= len(row):
                    continue
                metric_name = headers[mc]
                value_raw = row.iloc[mc]
                val_str = str(value_raw).strip()
                if val_str == '0':
                    cleaned_val = 0
                elif pd.isna(value_raw) or val_str.lower() in ['', '-', 'nan']:
                    cleaned_val = None
                else:
                    cleaned_val = val_str
                header_tuple = (table_heading_clean, metric_name, year_raw)
                processed_data.append({'Header': header_tuple, 'Value': cleaned_val})

    if debug:
        print(f"Processed {len(processed_data)} data points across {len(year_positions)} block(s)")
    return processed_data

# --- Special handling for combined Ph.D table (multiple mini-tables stacked) ---
def process_combined_phd_table(df_data: pd.DataFrame, table_heading: str, all_data: list, debug=False):
    """Splits the Ph.D combined table into two logical subtables and appends melted rows.

    Expected patterns:
      - One cell somewhere contains 'Total Students' (first mini-table metric heading)
      - A later row (not first) contains >=2 academic year patterns across non-first columns (year header row)
      - Rows between 'Total Students' header row and year header row contain category labels (Full Time, Part Time) with values under the 'Total Students' column
      - Rows after the year header row contain category labels with yearly values
    """
    if debug:
        print("Applying combined Ph.D table logic...")

    year_pattern = re.compile(r'20\d{2}-\d{2}')

    # Locate 'Total Students' cell
    total_students_pos = None  # (row_idx, col_idx)
    for r_idx in range(len(df_data)):
        for c_idx in range(len(df_data.columns)):
            cell = str(df_data.iat[r_idx, c_idx]).strip()
            if cell.lower() == 'total students':
                total_students_pos = (r_idx, c_idx)
                break
        if total_students_pos:
            break

    if not total_students_pos:
        if debug:
            print("Did not find 'Total Students' marker; falling back to default handling.")
        return False  # signal not processed

    ts_row, ts_col = total_students_pos

    # Find year header row
    year_header_row = None
    for r_idx in range(ts_row + 1, len(df_data)):
        row = df_data.iloc[r_idx]
        # Count year-like cells excluding first column
        year_hits = 0
        non_empty = 0
        for c_idx in range(1, len(df_data.columns)):
            val = str(row.iloc[c_idx]).strip()
            if val and val.lower() not in ['nan', '-', '']:
                non_empty += 1
                if year_pattern.fullmatch(val):
                    year_hits += 1
        if year_hits >= 2 and year_hits >= max(1, non_empty - year_hits):
            year_header_row = r_idx
            break

    if not year_header_row:
        if debug:
            print("Did not find year header row; falling back to default handling.")
        return False

    if debug:
        print(f"Found 'Total Students' at row {ts_row}, col {ts_col}; year header row at {year_header_row}")

    table_heading_clean = str(table_heading).replace('\n', ' ').strip()

    # --- Subtable 1: Total Students aggregate ---
    rows_added = 0
    for r_idx in range(ts_row + 1, year_header_row):
        category = str(df_data.iat[r_idx, 0]).replace('\n', ' ').strip()
        if not category or category.lower() in ['nan', '-', '']:
            continue
        raw_val = str(df_data.iat[r_idx, ts_col]).strip() if ts_col < len(df_data.columns) else ''
        if raw_val == '0':
            cleaned_val = 0
        elif raw_val.lower() in ['', '-', 'nan']:
            cleaned_val = None
        else:
            cleaned_val = raw_val
        header_tuple = (table_heading_clean, 'Total Students', category)
        all_data.append({'Header': header_tuple, 'Value': cleaned_val})
        rows_added += 1

    # --- Subtable 2: Yearly distribution ---
    years = []
    year_row = df_data.iloc[year_header_row]
    for c_idx in range(1, len(df_data.columns)):
        val = str(year_row.iloc[c_idx]).strip()
        if year_pattern.fullmatch(val):
            years.append((c_idx, val))

    for r_idx in range(year_header_row + 1, len(df_data)):
        category = str(df_data.iat[r_idx, 0]).replace('\n', ' ').strip()
        if not category or category.lower() in ['nan', '-', '']:
            continue
        for c_idx, year_label in years:
            if c_idx >= len(df_data.columns):
                continue
            value = str(df_data.iat[r_idx, c_idx]).strip()
            if value == '0':
                cleaned_val = 0
            elif value.lower() in ['', '-', 'nan']:
                cleaned_val = None
            else:
                cleaned_val = value
            header_tuple = (table_heading_clean, year_label, category)
            all_data.append({'Header': header_tuple, 'Value': cleaned_val})
            rows_added += 1

    if debug:
        print(f"Combined Ph.D table processed; added {rows_added} data points (aggregate + yearly).")
    return True  # processed

# --- Save to Excel (Main processing) ---

def process_and_save_to_excel(merged_tables, output_filename="master_output.xlsx", debug=False):
    """
    Processes extracted tables and saves them to a single master Excel file.
    """
    if not merged_tables:
        print("No tables to process for Excel."); return
    all_data = []
    print("\nProcessing tables for master Excel file...")
    
    processed_count = 0
    skipped_count = 0
    
    for table_heading, df in merged_tables.items():
        # Filter: only process tables with expenditure keywords in table data
        table_content = df.astype(str).values.flatten()
        table_text = ' '.join(table_content).lower()
        if not ("annual capital expenditure" in table_text or "annual operational expenditure" in table_text):
            if debug:
                print(f"\nSkipping table '{table_heading}' - does not contain required expenditure keywords in data")
            skipped_count += 1
            continue
            
        if debug:
            print(f"\n=== Processing table: '{table_heading}' ===")
            print(f"Original shape: {df.shape}")
        
        try:
            # Basic validation
            if df.shape[1] < 2 or df.shape[0] < 2:
                if debug: print(f"INFO: Skipping table '{table_heading}' due to insufficient size.")
                skipped_count += 1
                continue
                
            # Use the advanced function to find the header and data
            header_row_index, data_start_index, new_header = find_header_and_data_start(df)
            
            if debug:
                print(f"Header analysis results:")
                print(f"  Header row index: {header_row_index}")
                print(f"  Data start index: {data_start_index}")
                print(f"  New header: {new_header}")
                print(f"  Original shape: {df.shape}")
                if header_row_index < len(df):
                    print(f"  Header row content: {df.iloc[header_row_index].tolist()}")
                if data_start_index < len(df):
                    print(f"  First data row: {df.iloc[data_start_index].tolist()}")
            
            # Determine presence of specific columns for special handling
            lowered_headers = [str(h).lower() for h in new_header]
            contains_qualification = any("qualification" in h for h in lowered_headers)
            contains_designation = any("designation" in h for h in lowered_headers)
            contains_gender = any(h == "gender" for h in lowered_headers)

            # SPECIAL CASE: Employee roster table (has Qualification + Designation columns)
            # We don't want the raw 600+ row table. Instead, we aggregate counts by Designation and Gender
            if contains_qualification and contains_designation:
                if debug:
                    print(f"Detected employee roster table '{table_heading}'. Will aggregate counts by Designation and Gender instead of full melt.")

                # Validate data_start_index
                if data_start_index >= len(df):
                    if debug: print(f"WARNING: data_start_index {data_start_index} >= df length {len(df)} for roster table. Skipping.")
                    skipped_count += 1
                    continue

                df_data = df.iloc[data_start_index:].copy()

                # Assign headers (including de-dup if needed)
                if len(new_header) != len(df_data.columns):
                    if debug:
                        print(f"Roster table header length mismatch ({len(new_header)} vs {len(df_data.columns)}). Using original columns.")
                    new_header = df_data.columns.tolist()

                seen_headers = {}
                clean_headers = []
                for header in new_header:
                    if header in seen_headers:
                        seen_headers[header] += 1
                        clean_headers.append(f"{header}_{seen_headers[header]}")
                    else:
                        seen_headers[header] = 0
                        clean_headers.append(header)
                df_data.columns = clean_headers

                # Locate columns (case-insensitive)
                def find_col(target):
                    for c in df_data.columns:
                        if str(c).strip().lower() == target:
                            return c
                    return None

                designation_col = find_col('designation')
                gender_col = find_col('gender') if contains_gender else None

                if not designation_col:
                    if debug: print("ERROR: Could not find 'Designation' column after cleaning. Skipping roster aggregation.")
                    skipped_count += 1
                    continue

                table_heading_clean = str(table_heading).replace('\n', ' ').strip()
                rows_added = 0

                # Only aggregate gender counts (designation counts removed per request)
                if gender_col:
                    gender_series = (df_data[gender_col]
                                     .astype(str)
                                     .str.replace('\n', ' ', regex=False)
                                     .str.strip())
                    gender_series = gender_series[gender_series.str.len() > 0]
                    gender_counts = gender_series.value_counts(dropna=True)
                    for gender, count in gender_counts.items():
                        # Use blank third level to avoid adding a visible 'gender' title
                        header_tuple = (table_heading_clean, f"Number of {gender}", '')
                        all_data.append({'Header': header_tuple, 'Value': int(count)})
                        rows_added += 1

                if debug:
                    print(f"Added {rows_added} aggregated gender count columns from roster table (designation counts omitted).")
                processed_count += 1
                continue

            # Skip any other table that just has 'Qualification' (not the roster pattern we handle)
            if contains_qualification:
                if debug: print(f"INFO: Skipping table '{table_heading}' as it contains a 'Qualification' column (non-roster handling).")
                skipped_count += 1
                continue
            
            if debug:
                print(f"Header row index: {header_row_index}")
                print(f"Data start index: {data_start_index}")
                print(f"New header length: {len(new_header)}")
                print(f"New header: {new_header}")
            
            # Validate data_start_index
            if data_start_index >= len(df):
                if debug: print(f"WARNING: data_start_index {data_start_index} >= df length {len(df)}. Skipping.")
                skipped_count += 1
                continue
            
            # Special-case adjustment: for 2-column tables we want ALL rows treated as data.
            # These tables usually have no real header row; every row is a question/answer pair.
            if df.shape[1] == 2 and data_start_index > 0:
                if debug:
                    print(f"Overriding data_start_index {data_start_index} -> 0 for 2-column key/value table to avoid losing rows")
                data_start_index = 0

            # Use iloc for safer integer-based indexing
            df_data = df.iloc[data_start_index:].copy()
            
            # BEFORE proceeding, check if this is the combined Ph.D table pattern and process specially
            if 'ph.d (student pursuing doctoral program' in str(new_header[0]).lower() and df_data.shape[1] >= 3:
                if debug:
                    print("Attempting combined Ph.D table specialized split...")
                processed_ok = process_combined_phd_table(df_data, table_heading, all_data, debug=debug)
                if processed_ok:
                    processed_count += 1
                    continue
            
            if debug:
                print(f"Data frame shape after slicing: {df_data.shape}")
            
            if df_data.empty:
                if debug: print(f"INFO: Skipping table '{table_heading}' as no data rows found.")
                skipped_count += 1
                continue
            
            # Validate header length matches columns
            if len(new_header) != len(df_data.columns):
                if debug: 
                    print(f"WARNING: Header length {len(new_header)} != columns length {len(df_data.columns)}")
                    print(f"Using original headers instead.")
                new_header = df_data.columns.tolist()
            
            # Handle duplicate column names
            seen_headers = {}
            clean_headers = []
            for header in new_header:
                if header in seen_headers:
                    seen_headers[header] += 1
                    clean_headers.append(f"{header}_{seen_headers[header]}")
                else:
                    seen_headers[header] = 0
                    clean_headers.append(header)
            
            df_data.columns = clean_headers
            
            # Special handling: if the table has exactly 2 columns, treat
            # the first column as the header label and the second as the value.
            # We keep a 3-level tuple shape for consistency with other outputs
            # by using a constant placeholder (e.g. 'Value') for the middle level.
            if len(df_data.columns) == 2:
                if debug:
                    print("Detected 2-column table; applying key-value extraction logic")
                col_key, col_val = df_data.columns.tolist()
                rows_processed = 0
                table_heading_clean = str(table_heading).replace('\n', ' ').strip()
                for _, r in df_data.iterrows():
                    key_label = str(r[col_key]).replace('\n', ' ').strip()
                    val = r[col_val]
                    if not key_label or key_label.lower() in ['-', 'nan']:
                        continue
                    val_str = str(val).strip()
                    if val_str == '0':
                        cleaned_val = 0
                    elif pd.isna(val) or val_str.lower() in ['', '-', 'nan']:
                        cleaned_val = None
                    else:
                        cleaned_val = val_str
                    header_tuple = (table_heading_clean, key_label, '')  # third level blank
                    all_data.append({'Header': header_tuple, 'Value': cleaned_val})
                    rows_processed += 1
                if debug:
                    print(f"Successfully processed {rows_processed} key-value rows from 2-column table '{table_heading}'")
                processed_count += 1
                continue

            # Handle tables with repeating Academic Year columns (pivot structure)
            if clean_headers.count('Academic Year') > 1 or any('Academic Year_' in h for h in clean_headers):
                if debug: print(f"Detected pivot table structure with multiple Academic Year columns")
                processed_data = process_pivot_table(df_data, clean_headers, table_heading, debug)
                all_data.extend(processed_data)
                processed_count += 1
                continue
            
            # Validate we have at least one column for id_vars
            if len(df_data.columns) == 0:
                if debug: print(f"WARNING: No columns in table '{table_heading}'. Skipping.")
                skipped_count += 1
                continue
            
            id_col_name = clean_headers[0]
            
            # Validate identifier column
            if id_col_name == "" or pd.isna(id_col_name):
                if debug: print(f"WARNING: Skipping table '{table_heading}' due to empty identifier column.")
                skipped_count += 1
                continue
            
            if debug:
                print(f"Identifier column: '{id_col_name}'")
                print(f"All columns: {df_data.columns.tolist()}")
                print(f"First few rows of data:")
                print(df_data.head(2))
            
            # Perform the melt operation
            try:
                df_melted = df_data.melt(id_vars=[id_col_name], var_name='ColumnHeader', value_name='Value')
                
                if debug:
                    print(f"Melted data shape: {df_melted.shape}")
                    print(f"Melted data sample:")
                    print(df_melted.head(3))
                
            except Exception as melt_error:
                if debug: 
                    print(f"ERROR during melt operation: {melt_error}")
                    print(f"DataFrame info:")
                    print(f"  Shape: {df_data.shape}")
                    print(f"  Columns: {df_data.columns.tolist()}")
                    print(f"  id_col_name: {id_col_name}")
                skipped_count += 1
                continue
            
            # Process each row in the melted data
            rows_processed = 0
            for _, row in df_melted.iterrows():
                raw_val = row['Value']
                val_str = str(raw_val).strip()
                if val_str == '0':
                    cleaned_val = 0
                elif pd.isna(raw_val) or val_str.lower() in ['', '-', 'nan']:
                    cleaned_val = None
                else:
                    cleaned_val = raw_val
                table_heading_clean = str(table_heading).replace('\n', ' ').strip()
                row_category_clean = str(row[id_col_name]).replace('\n', ' ').strip()
                col_category_clean = str(row['ColumnHeader']).replace('\n', ' ').strip()
                header_tuple = (table_heading_clean, col_category_clean, row_category_clean)
                all_data.append({'Header': header_tuple, 'Value': cleaned_val})
                rows_processed += 1
            
            if debug: print(f"Successfully processed {rows_processed} rows from table '{table_heading}'")
            processed_count += 1
            
        except Exception as e:
            if debug: 
                print(f"ERROR processing table '{table_heading}': {e}")
                import traceback
                traceback.print_exc()
            skipped_count += 1
            continue
    
    print(f"\nProcessing summary: {processed_count} tables processed, {skipped_count} tables skipped")
    
    if not all_data:
        print("No valid data extracted to write to Excel."); return
        
    final_df = pd.DataFrame(all_data)
    
    if debug:
        print(f"\nFinal data shape: {final_df.shape}")
        print(f"Sample of final data:")
        print(final_df.head(10))
    
    # Proactively de-duplicate headers
    if final_df['Header'].duplicated().any():
        if debug: print("Found duplicate headers, applying de-duplication...")
        # Create a counter for each duplicated header
        counts = final_df.groupby('Header').cumcount()
        # Get the indices of the rows that are duplicates
        duplicated_indices = counts[counts > 0].index
        # Append a suffix to the last element of the tuple for each duplicate
        for idx in duplicated_indices:
            header_list = list(final_df.loc[idx, 'Header'])
            header_list[-1] = f"{header_list[-1]}_{counts[idx]}"
            final_df.loc[idx, 'Header'] = tuple(header_list)
    
    # Create multi-index for wide format
    multi_index = pd.MultiIndex.from_tuples(final_df['Header'])
    wide_df = pd.DataFrame([final_df['Value'].values], columns=multi_index)
    
    if debug:
        print(f"\nWide DataFrame shape: {wide_df.shape}")
        print(f"Wide DataFrame columns: {len(wide_df.columns)}")

    try:
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            wide_df.to_excel(writer, index=True, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            
            # Auto-fit columns
            for col_idx in range(1, worksheet.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                for cell in worksheet[col_letter]:
                    try:
                        if len(str(cell.value)) > max_length: 
                            max_length = len(str(cell.value))
                    except: 
                        pass
                adjusted_width = min((max_length + 2), 50)  # Cap at 50 characters
                worksheet.column_dimensions[col_letter].width = adjusted_width
        
        print(f"\n✅ Successfully saved master data to '{output_filename}'")
        print(f"✅ Total columns: {len(wide_df.columns)}")
        print(f"✅ File size: {os.path.getsize(output_filename) / 1024:.1f} KB")
        
    except Exception as e:
        print(f"\n❌ Error saving to master Excel file: {e}")
        import traceback
        traceback.print_exc()

# === Multi-PDF Aggregation Utilities (Folder Mode) ===

def extract_pdf_to_records(pdf_path: str, debug: bool=False) -> List[Dict]:
    """Extract one PDF and return list of {'Header': tuple, 'Value': value} records.
    This reuses the transformation logic of process_and_save_to_excel but stops before pivoting.
    Only processes tables with "Annual Capital Expenditure" or "Annual Operational Expenditure" in heading.
    Enhanced with year detection to improve data mapping.
    """
    import re
    
    merged_tables = extract_and_group_tables(pdf_path, debug=debug)
    if not merged_tables:
        return []
        
    records: List[Dict] = []
    
    # Year pattern to help identify academic years
    year_pattern = re.compile(r'(20\d{2}-\d{2}|20\d{2}-20\d{2})')
    
    # Debug information
    if debug:
        print(f"Processing {len(merged_tables)} tables for records extraction")
    
    # Minimal reimplementation: replicate internal logic but simpler (no wide pivot, no de-dupe across tables here)
    for table_heading, df in merged_tables.items():
        # Filter: only process tables with expenditure keywords in table data
        table_content = df.astype(str).values.flatten()
        table_text = ' '.join(table_content).lower()
        
        is_expenditure_table = ("annual capital expenditure" in table_text or 
                               "annual operational expenditure" in table_text or
                               "library" in table_text or "equipment" in table_text or 
                               "engineering workshop" in table_text or "salary" in table_text or
                               "seminars" in table_text)
        
        if not is_expenditure_table:
            if debug:
                print(f"Skipping non-expenditure table: {table_heading}")
            continue
            
        if debug:
            print(f"Processing potential expenditure table: {table_heading}")
            
        # Find any years mentioned in the table
        years_in_table = year_pattern.findall(table_text)
        if debug and years_in_table:
            print(f"  Found years in table: {years_in_table}")
            
        try:
            if df.shape[1] < 2 or df.shape[0] < 2:
                continue
            header_row_index, data_start_index, new_header = find_header_and_data_start(df)
            lowered_headers = [str(h).lower() for h in new_header]
            contains_qualification = any("qualification" in h for h in lowered_headers)
            contains_designation = any("designation" in h for h in lowered_headers)
            contains_gender = any(h == "gender" for h in lowered_headers)

            # Employee roster table -> gender aggregation only
            if contains_qualification and contains_designation:
                if data_start_index >= len(df):
                    continue
                df_data = df.iloc[data_start_index:].copy()
                if len(new_header) != len(df_data.columns):
                    new_header = df_data.columns.tolist()
                # Clean duplicate headers
                seen_headers = {}
                clean_headers = []
                for h in new_header:
                    if h in seen_headers:
                        seen_headers[h] += 1
                        clean_headers.append(f"{h}_{seen_headers[h]}")
                    else:
                        seen_headers[h] = 0
                        clean_headers.append(h)
                df_data.columns = clean_headers
                def find_col(target):
                    for c in df_data.columns:
                        if str(c).strip().lower() == target:
                            return c
                    return None
                gender_col = find_col('gender') if contains_gender else None
                if gender_col:
                    gender_series = (df_data[gender_col].astype(str).str.replace('\n', ' ', regex=False).str.strip())
                    gender_series = gender_series[gender_series.str.len() > 0]
                    for gender, count in gender_series.value_counts(dropna=True).items():
                        header_tuple = (table_heading, f"Number of {gender}", '')
                        records.append({'Header': header_tuple, 'Value': int(count)})
                continue

            if contains_qualification:
                continue

            if data_start_index >= len(df):
                continue

            # 2-col adjustment
            if df.shape[1] == 2 and data_start_index > 0:
                data_start_index = 0
            df_data = df.iloc[data_start_index:].copy()

            # Combined Ph.D special case
            if 'ph.d (student pursuing doctoral program' in str(new_header[0]).lower() and df_data.shape[1] >= 3:
                processed_ok = process_combined_phd_table(df_data, table_heading, records, debug=debug)
                if processed_ok:
                    continue

            if df_data.empty:
                continue
            if len(new_header) != len(df_data.columns):
                new_header = df_data.columns.tolist()
            seen_headers = {}
            clean_headers = []
            for h in new_header:
                if h in seen_headers:
                    seen_headers[h] += 1
                    clean_headers.append(f"{h}_{seen_headers[h]}")
                else:
                    seen_headers[h] = 0
                    clean_headers.append(h)
            df_data.columns = clean_headers

            # Key-value 2-column
            if len(df_data.columns) == 2:
                col_key, col_val = df_data.columns.tolist()
                for _, r in df_data.iterrows():
                    key_label = str(r[col_key]).replace('\n', ' ').strip()
                    if not key_label or key_label.lower() in ['-','nan']:
                        continue
                    val = r[col_val]
                    val_str = str(val).strip()
                    if val_str == '0':
                        cleaned_val = 0
                    elif pd.isna(val) or val_str.lower() in ['', '-', 'nan']:
                        cleaned_val = None
                    else:
                        cleaned_val = val_str
                    header_tuple = (table_heading, key_label, '')
                    records.append({'Header': header_tuple, 'Value': cleaned_val})
                continue

            # Pivot (repeating Academic Year)
            if clean_headers.count('Academic Year') > 1 or any(h.startswith('Academic Year_') for h in clean_headers):
                recs = process_pivot_table(df_data, clean_headers, table_heading, debug=False)
                records.extend(recs)
                continue

            # Check if any column header contains year patterns
            year_columns = []
            for header in clean_headers:
                if year_pattern.search(str(header)):
                    year_columns.append(header)
            
            if year_columns and debug:
                print(f"  Found column headers with years: {year_columns}")
            
            id_col = clean_headers[0]
            
            # Detect if we have a multi-year table with years in columns
            has_year_columns = len(year_columns) > 0
            
            # Process differently based on table structure
            if has_year_columns:
                # Handle table with years in columns - don't melt year columns together
                for year_col in year_columns:
                    year_match = year_pattern.search(str(year_col))
                    if year_match:
                        year_value = year_match.group(0)
                        for _, row in df_data.iterrows():
                            category = str(row[id_col]).replace('\n', ' ').strip()
                            if not category or category.lower() in ['-', 'nan', 'none']:
                                continue
                                
                            raw_val = row.get(year_col)
                            val_str = str(raw_val).strip()
                            if val_str == '0':
                                cleaned_val = 0
                            elif pd.isna(raw_val) or val_str.lower() in ['', '-', 'nan', 'none']:
                                cleaned_val = None
                            else:
                                cleaned_val = raw_val
                                
                            # Create header with year embedded in metric
                            metric = str(year_col).replace(year_value, '').strip()
                            if not metric:  # If year was the entire column header
                                metric = f"Value for {year_value}"
                                
                            header_tuple = (table_heading, f"{metric} ({year_value})", category)
                            records.append({'Header': header_tuple, 'Value': cleaned_val})
            else:
                # Standard approach for tables without year columns
                melted = df_data.melt(id_vars=[id_col], var_name='ColumnHeader', value_name='Value')
                for _, row in melted.iterrows():
                    raw_val = row['Value']
                    val_str = str(raw_val).strip()
                    if val_str == '0':
                        cleaned_val = 0
                    elif pd.isna(raw_val) or val_str.lower() in ['', '-', 'nan']:
                        cleaned_val = None
                    else:
                        cleaned_val = raw_val
                    row_cat = str(row[id_col]).replace('\n', ' ').strip()
                    col_cat = str(row['ColumnHeader']).replace('\n', ' ').strip()
                    
                    # Check for year in row category or column category
                    year_in_row = year_pattern.search(row_cat)
                    year_in_col = year_pattern.search(col_cat)
                    
                    # Adjust header tuple to include year information
                    if year_in_row:
                        year_value = year_in_row.group(0)
                        row_cat = row_cat.replace(year_value, '').strip()
                        header_tuple = (table_heading, f"{col_cat} ({year_value})", row_cat)
                    elif year_in_col:
                        year_value = year_in_col.group(0)
                        col_cat = col_cat.replace(year_value, '').strip()
                        header_tuple = (table_heading, f"{col_cat} ({year_value})", row_cat)
                    else:
                        header_tuple = (table_heading, col_cat, row_cat)
                        
                    records.append({'Header': header_tuple, 'Value': cleaned_val})
        except Exception:
            # Skip table on any unexpected error in folder mode to keep pipeline robust
            continue
    return records

def _normalize_header_drop_table(header_tuple: Tuple) -> Tuple[str, str]:
    """Drop the first element (table name) and reduce to (Main, Sub) tuple.
    If only one element after dropping, second level is blank.
    """
    if not header_tuple:
        return ("Unknown", "")
    # remove table heading
    parts = list(header_tuple)[1:]
    if not parts:
        return ("Unknown", "")
    if len(parts) == 1:
        return (parts[0], "")
    return (parts[0], parts[1])

def append_to_csv(pdf_records: List[Dict], pdf_filename: str, csv_filename: str = "out_format_new.csv"):
    """Append extracted data to CSV following the specified field mapping."""
    import csv
    import re
    from datetime import datetime
    
    # Define the CSV column headers and their corresponding data mapping
    csv_columns = [
        "College Name",
        "Library (2023-24)", "New Equipment for Laboratories (2023-24)", 
        "Engineering Workshops (2023-24)", "Studios (2023-24)", 
        "Other expenditure on creation of Capital Assets (2023-24)",
        "Library (2022-23)", "New Equipment for Laboratories (2022-23)",
        "Engineering Workshops (2022-23)", "Studios (2022-23)",
        "Other expenditure on creation of Capital Assets (2022-23)",
        "Library (2021-22)", "New Equipment for Laboratories (2021-22)",
        "Engineering Workshops (2021-22)", "Studios (2021-22)",
        "Other expenditure on creation of Capital Assets (2021-22)",
        "Salaries (2023-24)", "Maintenance of Academic Infrastructure or consumables and other running expenditures (2023-24)",
        "Seminars/Conferences/Workshops (2023-24)",
        "Salaries (2022-23)", "Maintenance of Academic Infrastructure or consumables and other running expenditures (2022-23)",
        "Seminars/Conferences/Workshops (2022-23)",
        "Salaries (2021-22)", "Maintenance of Academic Infrastructure or consumables and other running expenditures (2021-22)",
        "Seminars/Conferences/Workshops (2021-22)"
    ]
    
        # Create data mapping from extracted records
    data_map = {}
    debug_file = f"debug_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    
    with open(debug_file, "w", encoding="utf-8") as df:
        df.write(f"===== Processing {len(pdf_records)} records for CSV mapping =====\n\n")
        df.write(f"PDF Filename: {pdf_filename}\n\n")
        df.write(f"Script Version: Enhanced mapping with precise context v2.0\n\n")
        
        # First, let's see what records we have
        df.write("Sample records:\n")
        for i, rec in enumerate(pdf_records[:10]):
            df.write(f"  Record {i}: {rec}\n")
            
        # Add summary of expenditure keywords we're looking for
        df.write("\nMapping rules:\n")
        df.write("  Capital expenditure keywords: 'annual capital expenditure', 'library expenditure', 'laboratory equipment', etc.\n")
        df.write("  Operational expenditure keywords: 'annual operational expenditure', 'salary expenditure', etc.\n")
        df.write("  Excluded patterns: 'median salary', 'placed graduates', etc.\n\n")        # Year pattern for more robust extraction
        year_pattern = re.compile(r'(20\d{2}-\d{2}|20\d{2}-20\d{2})')
        
        for rec in pdf_records:
            header_tuple = rec['Header']
            value = rec['Value']
            
            if value is None or str(value).strip() == '':
                continue
                
            # Convert header tuple parts to strings and clean
            table_heading = str(header_tuple[0]).strip().lower() if len(header_tuple) > 0 else ""
            metric = str(header_tuple[1]).strip().lower() if len(header_tuple) > 1 else ""
            category = str(header_tuple[2]).strip().lower() if len(header_tuple) > 2 else ""
            
            # Special case handling for complex headers with expenditure information
            capital_asset_pattern = re.compile(r'other expenditure on creation of capital assets.*\(for setting up|excluding.*land', re.IGNORECASE)
            
            # Check if any part of the header contains the complex pattern
            complex_capital_asset = False
            for part in header_tuple:
                if isinstance(part, str) and capital_asset_pattern.search(part.lower()):
                    complex_capital_asset = True
                    df.write(f"  Found complex capital asset pattern in header\n")
                    break
            
            # Combine all text for better keyword detection
            full_content = f"{table_heading} {metric} {category}".lower()
            df.write(f"\nProcessing record: {header_tuple} = {value}\n")
            df.write(f"  Combined content: '{full_content}'\n")
            
            # If complex capital asset pattern was detected, add a flag
            if complex_capital_asset:
                df.write(f"  SPECIAL CASE: Complex capital asset header detected\n")
                
            # Special case for library-related fields
            library_pattern = re.compile(r'library\s*\(\s*books', re.IGNORECASE)
            is_library_record = False
            is_studio_record = False
            
            # Check if any part of the header contains the library pattern
            for part in header_tuple:
                if isinstance(part, str) and library_pattern.search(part.lower()):
                    is_library_record = True
                    df.write(f"  SPECIAL CASE: Library with books/journals detected\n")
                    break
                    
            # Also check if any part starts with "Library"
            if not is_library_record:
                for part in header_tuple:
                    if isinstance(part, str) and part.lower().strip().startswith("library"):
                        is_library_record = True
                        
            # Check if any part is "Studios"
            for part in header_tuple:
                if isinstance(part, str) and part.lower().strip() == "studios":
                    is_studio_record = True
                    df.write(f"  SPECIAL CASE: Studios detected as capital expenditure\n")
                    break
            
            # Check if the record is expenditure data - use more specific context-aware checks
            # First, filter out non-expenditure related data
            if "median salary" in full_content:
                df.write("  Skipping - median salary is not institutional expenditure\n")
                continue
                
            if "placed graduates" in full_content:
                df.write("  Skipping - placement information is not expenditure data\n")
                continue
                
            # Check for specific expenditure keywords with better context
            is_capital_data = False
            is_operational_data = False
            
            # Capital expenditure checks with context
            if "annual capital expenditure" in full_content:
                is_capital_data = True
            elif "library expenditure" in full_content or "expenditure on library" in full_content:
                is_capital_data = True
            elif is_library_record:  # Use our special case detection
                is_capital_data = True
                df.write("  Identified as library expenditure data based on header pattern\n")
            elif is_studio_record or "studios" in full_content:  # Use our special case detection for studios
                is_capital_data = True
                df.write("  Identified as studio expenditure data\n")
            elif ("equipment" in full_content and "laborator" in full_content) or "laboratory equipment" in full_content:
                is_capital_data = True
            elif "engineering workshop" in full_content and not "seminar" in full_content:
                is_capital_data = True
            elif "studio" in full_content and not any(non_exp in full_content for non_exp in ["student", "graduating", "placed"]):
                is_capital_data = True
            elif "capital asset" in full_content and not "other" in full_content:
                is_capital_data = True
            
            # First check for specific header patterns that should always be recognized
            is_salary_record = False
            is_seminar_record = False
            is_maintenance_record = False
            is_studio_record = False
            
            # Check if this is a salary record by its header
            if "salaries (teaching and non teaching staff)" in full_content:
                is_salary_record = True
                is_operational_data = True
                df.write("  Identified as salary data based on header pattern\n")
            
            # Check if this is a seminar/conference record by its header
            elif "seminars/conferences/workshops" in full_content:
                is_seminar_record = True
                is_operational_data = True
                df.write("  Identified as seminar/conference data based on header pattern\n")
                
            # Operational expenditure checks with context
            elif "annual operational expenditure" in full_content:
                is_operational_data = True
            elif ("salary expenditure" in full_content or 
                  ("salary" in full_content and "expenditure" in full_content) or
                  ("salaries" in full_content and "teaching" in full_content and "non teaching" in full_content)):
                # Exclude median salary of placed graduates
                if not any(excl in full_content for excl in ["median", "placed", "graduate"]):
                    is_salary_record = True
                    is_operational_data = True
            elif "maintenance" in full_content and "infrastructure" in full_content:
                is_maintenance_record = True
                is_operational_data = True
            elif ("seminar/conference/workshop" in full_content or
                  ("seminar" in full_content or "conference" in full_content or "workshop" in full_content) and "expenditure" in full_content):
                is_seminar_record = True
                is_operational_data = True
                
            # Handle special case for "Other expenditure" which needs specific categorization
            if "other expenditure" in full_content:
                if "capital" in full_content or "asset" in full_content:
                    is_capital_data = True
                    is_operational_data = False
            
            df.write(f"  Is Capital: {is_capital_data}, Is Operational: {is_operational_data}\n")
            
            if not (is_capital_data or is_operational_data):
                df.write("  Skipping - not expenditure data\n")
                continue
            
            # Extract year from any part of the header tuple
            # First, try direct year pattern match
            year_match = year_pattern.search(full_content)
            year_found = year_match.group(0) if year_match else None
            
            # If no direct year match, try to infer from context or related records
            if not year_found:
                for year in ["2023-24", "2022-23", "2021-22"]:
                    # Check if there are indicators of the year in nearby context
                    for related_rec in pdf_records:
                        related_header = related_rec['Header']
                        if len(related_header) >= 2:
                            related_content = " ".join([str(part).strip().lower() for part in related_header])
                            if year in related_content and any(kw in full_content for kw in [
                                metric, category, table_heading
                            ]):
                                year_found = year
                                df.write(f"  Year inferred from related record: {year}\n")
                                break
                    if year_found:
                        break
            
            # Still no year? Try to infer from current/previous/next academic year context
            if not year_found:
                current_year = datetime.now().year
                if f"current" in full_content or f"{current_year}" in full_content:
                    year_found = "2023-24"  # Assume current is 2023-24
                    df.write(f"  Year inferred from 'current year' context: {year_found}\n")
                elif f"previous" in full_content or f"{current_year-1}" in full_content:
                    year_found = "2022-23"  # Assume previous is 2022-23
                    df.write(f"  Year inferred from 'previous year' context: {year_found}\n")
                elif f"before" in full_content or f"{current_year-2}" in full_content:
                    year_found = "2021-22"  # Assume year before previous is 2021-22
                    df.write(f"  Year inferred from 'year before previous' context: {year_found}\n")
            
            if not year_found:
                df.write("  No year found, trying to match based on column position\n")
                # Try to identify the column/row position to guess the year
                if "3" in metric or "3rd" in metric or "third" in metric:
                    year_found = "2021-22"
                elif "2" in metric or "2nd" in metric or "second" in metric:
                    year_found = "2022-23"
                elif "1" in metric or "1st" in metric or "first" in metric:
                    year_found = "2023-24"
            
            if not year_found:
                df.write("  Skipping - could not determine year\n")
                continue
                
            df.write(f"  Year found/inferred: {year_found}\n")
            
            # Map to CSV columns based on data content and year with more precise matching
            mapped = False
            
            if is_capital_data:
                # Map capital expenditure fields with more precise context matching
                
                # Library - priority check for our special library record detection
                if is_library_record:
                    column_key = f"Library ({year_found})"
                    data_map[column_key] = value
                    df.write(f"  MAPPED to {column_key}: {value}\n")
                    df.write(f"  Mapped based on library header pattern detection\n")
                    mapped = True
                # Other library patterns
                elif ("library expenditure" in full_content or 
                    "expenditure on library" in full_content or 
                    "library ( books" in full_content or  # Specifically match "Library ( Books, Journals and e-Resources only)"
                    category.startswith("library") or     # Match category field starting with "library"
                    metric.startswith("library") or       # Match metric field starting with "library" 
                    (("library" in full_content or "libraries" in full_content) and 
                     ("expenditure" in full_content or "books" in full_content or "journals" in full_content or "e-resources" in full_content) and 
                     not any(excl in full_content for excl in ["equipment", "lab", "workshop", "studio", "other"]))):
                    column_key = f"Library ({year_found})"
                    data_map[column_key] = value
                    df.write(f"  MAPPED to {column_key}: {value}\n")
                    mapped = True
                    
                # Laboratory equipment
                elif (("equipment" in full_content and "laborator" in full_content) or 
                      "laboratory equipment" in full_content or 
                      ("new equipment" in full_content and not "other" in full_content)):
                    column_key = f"New Equipment for Laboratories ({year_found})"
                    data_map[column_key] = value
                    df.write(f"  MAPPED to {column_key}: {value}\n")
                    mapped = True
                    
                # Engineering Workshops
                elif ("engineering workshop" in full_content or 
                      ("workshop" in full_content and "engineering" in full_content) and 
                      not any(excl in full_content for excl in ["seminar", "conference"])):
                    column_key = f"Engineering Workshops ({year_found})"
                    data_map[column_key] = value
                    df.write(f"  MAPPED to {column_key}: {value}\n")
                    mapped = True
                    
                # Studios - use special case detection or fallback to keyword matching
                elif is_studio_record or "studio" in full_content or "studios" in full_content:
                    column_key = f"Studios ({year_found})"
                    data_map[column_key] = value
                    df.write(f"  MAPPED to {column_key}: {value}\n")
                    mapped = True
                    
                # Other capital expenditure
                elif (("other expenditure" in full_content and "capital" in full_content) or
                      ("creation of capital asset" in full_content) or
                      ("other" in full_content and "capital asset" in full_content)):
                    column_key = f"Other expenditure on creation of Capital Assets ({year_found})"
                    data_map[column_key] = value
                    df.write(f"  MAPPED to {column_key}: {value}\n")
                    mapped = True
                        
            elif is_operational_data:
                # Map operational expenditure fields
                
                # Salaries - prioritize direct header detection
                if is_salary_record or (("salary expenditure" in full_content or "salaries expenditure" in full_content) or
                    "salaries (teaching and non teaching staff)" in full_content or
                    "teaching and non teaching staff" in full_content or
                    ("teaching staff" in full_content and "non teaching staff" in full_content) or
                    (("salary" in full_content or "salaries" in full_content) and 
                     ("expenditure" in full_content or "value" in full_content) and 
                     not any(excl in full_content for excl in ["median", "placed", "graduate"]))):
                    column_key = f"Salaries ({year_found})"
                    data_map[column_key] = value
                    df.write(f"  MAPPED to {column_key}: {value}\n")
                    mapped = True
                    
                # Maintenance and infrastructure
                elif (("maintenance" in full_content and "infrastructure" in full_content) or
                      "academic infrastructure" in full_content or 
                      "consumable" in full_content and "expenditure" in full_content):
                    column_key = f"Maintenance of Academic Infrastructure or consumables and other running expenditures ({year_found})"
                    data_map[column_key] = value
                    df.write(f"  MAPPED to {column_key}: {value}\n")
                    mapped = True
                    
                # Seminars/Conferences/Workshops - prioritize direct header detection
                elif (is_seminar_record or 
                     ((any(term in full_content for term in ["seminar", "conference", "workshop", "seminar/conference/workshop"]) and 
                       ("expenditure" in full_content or "expenditure on" in full_content or "value" in full_content)) and
                      not "engineering workshop" in full_content) or
                      "seminars/conferences/workshops" in full_content):
                    column_key = f"Seminars/Conferences/Workshops ({year_found})"
                    data_map[column_key] = value
                    df.write(f"  MAPPED to {column_key}: {value}\n")
                    mapped = True
                    
            # Handle the specific case mentioned in the user's example - a combined heading with multiple types
            if not mapped:
                # Check for complex capital asset pattern
                if complex_capital_asset or "other expenditure on creation of capital assets" in full_content:
                    # Different variations of the pattern
                    if ("(for setting up" in full_content or 
                        "excluding expenditure on land" in full_content or
                        ("other" in full_content and "capital" in full_content and "asset" in full_content)):
                        # This is a combined entry that should go to "Other expenditure on creation of Capital Assets"
                        column_key = f"Other expenditure on creation of Capital Assets ({year_found})"
                        data_map[column_key] = value
                        df.write(f"  MAPPED to {column_key}: {value}\n")
                        mapped = True
                        df.write(f"  NOTE: Complex capital asset pattern recognized and properly mapped\n")
                    
            if not mapped:
                df.write(f"  FAILED TO MAP - could not determine appropriate column for content: {full_content}\n")
                
        # Prepare the row data
        college_name = pdf_filename.replace('.pdf', '').replace('_', ' ').title()
        row_data = [college_name]
        
        # Add data for each column (skip College Name which is first)
        for col in csv_columns[1:]:
            row_data.append(data_map.get(col, ""))
        
        df.write(f"\nFinal mapped data for {college_name}:\n")
        for i, col in enumerate(csv_columns):
            if i == 0:
                df.write(f"  {col}: {college_name}\n")
            else:
                df.write(f"  {col}: {row_data[i] if i < len(row_data) else ''}\n")
        
        print(f"Processed {pdf_filename} - see {debug_file} for detailed mapping information")
    
    # Create CSV file with headers if it doesn't exist
    if not os.path.exists(csv_filename):
        try:
            with open(csv_filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(csv_columns)
            print(f"Created new CSV file: {csv_filename}")
        except Exception as e:
            print(f"❌ Error creating CSV file: {e}")
            return
    
    # Append to CSV file
    try:
        with open(csv_filename, 'a', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(row_data)
        print(f"✅ Successfully appended data for {college_name} to {csv_filename}")
        
        # Also save mapping details in debug folder
        debug_folder = "mapping_debug"
        if not os.path.exists(debug_folder):
            os.makedirs(debug_folder)
        
        debug_csv = os.path.join(debug_folder, f"{college_name.replace(' ', '_')}_mapping.csv")
        with open(debug_csv, 'w', newline='', encoding='utf-8') as mapfile:
            map_writer = csv.writer(mapfile)
            # Write header
            map_writer.writerow(["Column", "Value"])
            # Write college name and all mapped values
            map_writer.writerow(["College Name", college_name])
            for col in csv_columns[1:]:
                map_writer.writerow([col, data_map.get(col, "")])
        
    except Exception as e:
        print(f"❌ Error appending to CSV: {e}")

def process_folder(input_folder: str, output_filename: str="master_output.xlsx", debug: bool=False):
    """Process all PDFs in a folder and aggregate into a single master_output.xlsx.
    - Ignores table names: columns defined purely by (Metric, Category)
    - Each PDF becomes one row in the Excel sheet (index = PDF filename)
    - Matching headings across PDFs share the same column; new headings create new columns.
    - Only processes tables with "Annual Capital Expenditure" or "Annual Operational Expenditure" in heading.
    """
    if not os.path.isdir(input_folder):
        print(f"Input folder '{input_folder}' not found.")
        return
    pdf_files = [f for f in os.listdir(input_folder) if f.lower().endswith('.pdf')]
    if not pdf_files:
        print(f"No PDF files found in '{input_folder}'.")
        return
    pdf_files.sort()
    
    # Create CSV filename based on folder name
    folder_name = os.path.basename(input_folder.rstrip('/\\'))
    csv_filename = f"{folder_name.replace(' ', '_')}_output.csv"
    
    print(f"Found {len(pdf_files)} PDF(s) in folder '{input_folder}'.")
    print(f"CSV output will be saved as: {csv_filename}")
    global_columns: List[Tuple[str,str]] = []
    row_data: List[Dict] = []
    for pdf_name in pdf_files:
        pdf_path = os.path.join(input_folder, pdf_name)
        print(f"\n--- Processing PDF: {pdf_name} ---")
        
        # Extract tables and save debug info for each PDF
        merged_tables = extract_and_group_tables(pdf_path, debug=debug)
        if merged_tables:
            save_detailed_debug_info(merged_tables, pdf_path)
        
        recs = extract_pdf_to_records(pdf_path, debug=debug)
        
        # Save records debug info for each PDF
        save_records_debug(recs, pdf_name)
        
        # Append to CSV file with folder-specific name
        append_to_csv(recs, pdf_name, csv_filename)
        
        normalized_map: Dict[Tuple[str,str], object] = {}
        for rec in recs:
            # No additional filtering needed here since extract_pdf_to_records already filters by table data
            norm = _normalize_header_drop_table(rec['Header'])
            value = rec['Value']
            if norm not in normalized_map:
                normalized_map[norm] = value
            else:
                # Merge rule: keep first non-null; if existing is None and new is not None, replace
                if (normalized_map[norm] is None) and (value is not None):
                    normalized_map[norm] = value
            if norm not in global_columns:
                global_columns.append(norm)
        row_data.append({'__pdf__': pdf_name, '__map__': normalized_map})

    # Build DataFrame (MultiIndex columns)
    multi_index = pd.MultiIndex.from_tuples(global_columns, names=['Heading','SubHeading'])
    matrix = []
    idx = []
    for rd in row_data:
        row_values = [rd['__map__'].get(col, None) for col in global_columns]
        matrix.append(row_values)
        idx.append(rd['__pdf__'])
    wide_df = pd.DataFrame(matrix, columns=multi_index, index=idx)

    # Flatten columns to single-level strings in the format "SubHeading (Heading)"
    def _fmt_col(col_tuple: tuple) -> str:
        try:
            head, sub = col_tuple
        except Exception:
            # Fallback if not a pair
            return str(col_tuple)
        sub_s = str(sub).strip()
        head_s = str(head).strip()
        # If SubHeading is empty, use just Heading to avoid blank prefixes
        return f"{sub_s} ({head_s})" if sub_s else head_s

    flat_cols = [_fmt_col(c) for c in wide_df.columns]
    wide_df.columns = flat_cols

    try:
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            wide_df.to_excel(writer, sheet_name='Sheet1')
            ws = writer.sheets['Sheet1']
            # Autosize
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for cell in ws[col_letter]:
                    try:
                        if cell.value is not None:
                            max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
        print(f"\n✅ Aggregated {len(pdf_files)} PDFs into '{output_filename}' with {len(global_columns)} unified columns.")
    except Exception as e:
        print(f"Failed to write aggregated Excel: {e}")

# --- Main ---

def main(debug=False, input_path: Optional[str]=None):
    """Entry point supporting two modes:
    - Folder mode: if 'input files' (default) or provided input_path is a directory, process all PDFs and aggregate.
    - Single PDF mode: fallback to original behavior for a lone PDF file.
    """
    folder_candidate = "NIRF IIT'S"
    if os.path.isdir(folder_candidate):
        print(f"Running in multi-PDF aggregation mode for folder: {folder_candidate}")
        process_folder(folder_candidate, output_filename="master_output.xlsx", debug=debug)
        return
    # Single PDF fallback
    pdf_path = input_path or 'iit_delhi_data.pdf'
    if not os.path.exists(pdf_path):
        print(f"Error: PDF file not found at '{pdf_path}'.")
        return
    print("\nStarting table extraction and grouping (single PDF mode)...")
    merged_tables = extract_and_group_tables(pdf_path, debug=debug)
    if merged_tables:
        save_raw_tables_for_debug(merged_tables)
        
        # Save detailed debugging information
        save_detailed_debug_info(merged_tables, pdf_path)
        
        # Extract records and append to CSV
        recs = extract_pdf_to_records(pdf_path, debug=debug)
        pdf_filename = os.path.basename(pdf_path)
        
        # Save records to debug file
        save_records_debug(recs, pdf_filename)
        
        # For single PDF, use the PDF name as CSV filename
        single_csv_name = f"{pdf_filename.replace('.pdf', '')}_output.csv"
        append_to_csv(recs, pdf_filename, single_csv_name)
        
        process_and_save_to_excel(merged_tables, output_filename="master_output.xlsx", debug=debug)
    else:
        print("No tables were extracted from the PDF.")

def save_detailed_debug_info(merged_tables, pdf_path):
    """Save detailed debugging information about extracted tables in organized debug folder"""
    # Create debug folder if it doesn't exist
    debug_folder = "extracted_tables_debug"
    if not os.path.exists(debug_folder):
        os.makedirs(debug_folder)
    
    # Get clean PDF filename without extension
    pdf_name = os.path.splitext(os.path.basename(pdf_path))[0]
    debug_filename = os.path.join(debug_folder, f"{pdf_name}_extracted_tables.txt")
    
    with open(debug_filename, 'w', encoding='utf-8') as f:
        f.write(f"EXTRACTED TABLE DATA FOR: {pdf_name}\n")
        f.write(f"Source PDF: {pdf_path}\n")
        f.write(f"Generated at: {datetime.now()}\n")
        f.write("="*80 + "\n\n")
        
        for table_name, df in merged_tables.items():
            f.write(f"TABLE: {table_name}\n")
            f.write(f"Shape: {df.shape} (Rows: {df.shape[0]}, Columns: {df.shape[1]})\n")
            f.write(f"Column Names: {list(df.columns)}\n")
            f.write("-" * 60 + "\n")
            f.write("RAW TABLE DATA:\n")
            f.write(df.to_string(index=True, max_rows=None, max_cols=None))
            f.write("\n\n")
            
            # Also save as CSV for easier analysis
            csv_filename = os.path.join(debug_folder, f"{pdf_name}_{table_name.replace(' ', '_').replace('/', '_')}.csv")
            try:
                df.to_csv(csv_filename, index=False, encoding='utf-8')
                f.write(f"Also saved as CSV: {csv_filename}\n")
            except Exception as e:
                f.write(f"Failed to save CSV: {e}\n")
            
            f.write("\n" + "="*80 + "\n\n")
    
    log_print(f"Debug info for {pdf_name} saved to: {debug_filename}")

def save_records_debug(records, pdf_filename):
    """Save extracted records for debugging in organized debug folder"""
    # Create debug folder if it doesn't exist
    debug_folder = "extracted_tables_debug"
    if not os.path.exists(debug_folder):
        os.makedirs(debug_folder)
    
    pdf_name = os.path.splitext(pdf_filename)[0]
    records_filename = os.path.join(debug_folder, f"{pdf_name}_processed_records.txt")
    
    with open(records_filename, 'w', encoding='utf-8') as f:
        f.write(f"PROCESSED RECORDS FOR: {pdf_name}\n")
        f.write(f"Generated at: {datetime.now()}\n")
        f.write(f"Total Records: {len(records)}\n")
        f.write("="*80 + "\n\n")
        
        for i, record in enumerate(records):
            f.write(f"Record {i+1}:\n")
            f.write(f"  Header Tuple: {record.get('Header', 'N/A')}\n")
            f.write(f"  Value: {record.get('Value', 'N/A')}\n")
            f.write("-" * 50 + "\n")
    
    log_print(f"Processed records for {pdf_name} saved to: {records_filename}")

if __name__ == "__main__":
    # By default try folder mode; set debug=False for cleaner output when running on many PDFs
    main(debug=True)  # Changed to True for detailed debugging